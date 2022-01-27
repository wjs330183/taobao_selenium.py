import openpyxl
from selenium import webdriver  # 用来驱动浏览器的
import time




def getExcelDetail():
    readbook = openpyxl.load_workbook('C:\\Users\\administor\\Documents\\zb.xlsx')
    # 名字的方式
    sheetnames = readbook.get_sheet_names()  # 获取读文件中所有的sheet，通过名字的方式
    ws = readbook.get_sheet_by_name(sheetnames[0])  # 获取第一个sheet内容
    # 获取sheet的最大行数和列数
    rows = ws.max_row
    cols = ws.max_column + 1
    for r in range(1, rows):
        for c in range(1, cols):
            print(ws.cell(r, c).value)

'''
===============所有方法===================
    element是查找一个标签
    elements是查找所有标签

    1、find_element_by_link_text  通过链接文本去找
    2、find_element_by_id 通过id去找
    3、find_element_by_class_name
    4、find_element_by_partial_link_text
    5、find_element_by_name
    6、find_element_by_css_selector
    7、find_element_by_tag_name
'''
# 获取驱动对象、
driver = webdriver.Chrome()

try:

    # 往百度发送请求
    driver.get('https://www.baidu.com/')
    driver.implicitly_wait(10)

    # 1、find_element_by_link_text  通过链接文本去找
    # 根据登录
    # send_tag = driver.find_eleme
    #
    #
    #
    #
    # nt_by_link_text('登录')
    # send_tag.click()

    # 2、find_element_by_partial_link_text 通过局部文本查找a标签
    login_button = driver.find_element_by_partial_link_text('登')
    login_button.click()
    time.sleep(1)

    # 3、find_element_by_class_name 根据class属性名查找
    # login_tag = driver.find_element_by_class_name('tang-pass-footerBarULogin')
    # login_tag.click()
    # time.sleep(1)

    # 4、find_element_by_name 根据name属性查找
    username = driver.find_element_by_name('userName')
    username.send_keys('18758065923')
    time.sleep(1)

    # 5、find_element_by_id 通过id属性名查找
    password = driver.find_element_by_id('TANGRAM__PSP_11__password')
    password.send_keys('wang78952')
    time.sleep(1)

    # 6、find_element_by_css_selector  根据属性选择器查找
    # 根据id查找登录按钮
    login_submit = driver.find_element_by_css_selector('#TANGRAM__PSP_11__submit')
    driver.find_element_by_css_selector('.pass-button-submit')
    login_submit.click()

    # 7、find_element_by_tag_name  根据标签名称查找标签
    div = driver.find_element_by_tag_name('div')
    print(div.tag_name)

    time.sleep(10)

finally:
    driver.close()