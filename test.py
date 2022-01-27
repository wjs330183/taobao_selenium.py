import urllib.request

import openpyxl
import requests
import json
import re

import xlrd


def print_hi(name):
    res = urllib.request.urlopen("https://www.qcc.com/cnews/bc80e695ea8504c3d2bed9cc5250ef03.html")
    r = requests.get("https://www.qcc.com/cnews/bc80e695ea8504c3d2bed9cc5250ef03.html")
    text = r.text
    status_code = r.status_code
    # json = r.json()
    cookies = r.cookies
    headers = r.headers
    rheaders = r.request.headers
    # print(r.text)
    print(rheaders)
    print(headers)
    print(cookies)
    print(r.status_code)
    # print(r.json())
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.


def getExcelDetail():
    readbook = openpyxl.load_workbook('C:\\Users\\administor\\Documents\\zb.xlsx')
    # 名字的方式
    sheetnames = readbook.get_sheet_names()  # 获取读文件中所有的sheet，通过名字的方式
    ws = readbook.get_sheet_by_name(sheetnames[0])  # 获取第一个sheet内容
    # 获取sheet的最大行数和列数
    rows = ws.max_row
    cols = ws.max_column + 1
    list = [[0 for i in range(ws.max_column + 1)] for j in range(ws.max_row)]
    name = 'None'
    local = '['
    for r in range(1, rows):
        if (r > 1):
            local = local + ','
        local = local + '['
        for c in range(1, cols):
            if ('None' != str(ws.cell(r, 1).value) and 1 == c):
                local = local[0:-2]
                if (local != ""):
                    local = local + ']'
                    name = seaName
                    seaNumber = name

                    print(name + "      " + local)
                local = '[['
                seaName = ws.cell(r, 1).value
                seaNumber = seaName
            elif (1 != c):
                value = ws.cell(r, c).value
                sb = str(value).split('°', 1)
                sb1 = sb[1].split('′', 1)
                value1 = sb[0]
                value2 = sb1[0]
                value3 = sb1[1]
                value3 = value3[0:-1]

                local = local + value
                if (2 == c):
                    local = local + ','

        local = local + ']'

    print("`end`")


if __name__ == '__main__':
    # print_hi('PyCharm')
    # getExcelDetail()
    title = '123-123-123'
    titles = title.split("-", 1)
    len(titles)
    print(len(titles))