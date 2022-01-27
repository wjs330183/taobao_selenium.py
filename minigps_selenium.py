import openpyxl
from selenium import webdriver  # 用来驱动浏览器的
import time

from pyquery import PyQuery as pq
from peewee import *
import time
import datetime

db = MySQLDatabase("g_dt_test", host="rm-bp10hal2887x18nb97o.mysql.rds.aliyuncs.com", port=3306, user="g_dt_test",
                   passwd="Palcomm007")


class SeaPlan(Model):
    id = CharField()
    sea_number = CharField()
    cate_code = CharField()
    area_type = CharField()
    water_type = CharField()
    cultivate_type = CharField()
    acreage = DecimalField()
    coord = CharField()
    principal = CharField()
    town = CharField()
    color = CharField()
    sea_type = CharField()
    parent_id = CharField()
    idle_acreage = DecimalField()
    comment = CharField()
    create_by = CharField()
    create_time = DateTimeField()
    update_by = CharField()
    update_time = DateTimeField()

    class Meta:
        database = db
        table_name = 'sea_plan'


def save(seaNumber, cateCode, areaType, waterType, cultivateType, acreage, coord,
         principal, town, color, seaType, parentId, idleAcreage, comment):
    worker = IdWorker(0, 0)
    id = worker.get_id()
    createBy = "admin"
    createTime = datetime.datetime.now()
    updateBy = "admin"
    updateTime = datetime.datetime.now()

    seaPlan = SeaPlan(id=id,
                      sea_number=seaNumber,
                      cate_code=cateCode,
                      area_type=areaType,
                      water_type=waterType,
                      cultivate_type=cultivateType,
                      acreage=acreage,
                      coord=coord,
                      principal=principal,
                      town=town,
                      color=color,
                      sea_type=seaType,
                      parent_id=parentId,
                      idle_acreage=idleAcreage,
                      comment=comment,
                      create_by=createBy,
                      create_time=createTime,
                      update_by=updateBy,
                      update_time=updateTime)
    seaPlan.save(force_insert=True)


class InvalidSystemClock(Exception):
    """
    时钟回拨异常
    """
    pass


# 64位ID的划分
WORKER_ID_BITS = 5
DATACENTER_ID_BITS = 5
SEQUENCE_BITS = 12

# 最大取值计算
MAX_WORKER_ID = -1 ^ (-1 << WORKER_ID_BITS)  # 2**5-1 0b11111
MAX_DATACENTER_ID = -1 ^ (-1 << DATACENTER_ID_BITS)

# 移位偏移计算
WOKER_ID_SHIFT = SEQUENCE_BITS
DATACENTER_ID_SHIFT = SEQUENCE_BITS + WORKER_ID_BITS
TIMESTAMP_LEFT_SHIFT = SEQUENCE_BITS + WORKER_ID_BITS + DATACENTER_ID_BITS

# 序号循环掩码
SEQUENCE_MASK = -1 ^ (-1 << SEQUENCE_BITS)

# 开始时间截 (2015-01-01)
TWEPOCH = 1420041600000


class IdWorker(object):
    """
    用于生成IDs
    """

    def __init__(self, datacenter_id, worker_id, sequence=0):
        """
        初始化
        :param datacenter_id: 数据中心（机器区域）ID
        :param worker_id: 机器ID
        :param sequence: 其实序号
        """
        # sanity check
        if worker_id > MAX_WORKER_ID or worker_id < 0:
            raise ValueError('worker_id值越界')

        if datacenter_id > MAX_DATACENTER_ID or datacenter_id < 0:
            raise ValueError('datacenter_id值越界')

        self.worker_id = worker_id
        self.datacenter_id = datacenter_id
        self.sequence = sequence

        self.last_timestamp = -1  # 上次计算的时间戳

    def _gen_timestamp(self):
        """
        生成整数时间戳
        :return:int timestamp
        """
        return int(time.time() * 1000)

    def get_id(self):
        """
        获取新ID
        :return:
        """
        timestamp = self._gen_timestamp()

        # 时钟回拨
        if timestamp < self.last_timestamp:
            raise InvalidSystemClock

        if timestamp == self.last_timestamp:
            self.sequence = (self.sequence + 1) & SEQUENCE_MASK
            if self.sequence == 0:
                timestamp = self._til_next_millis(self.last_timestamp)
        else:
            self.sequence = 0

        self.last_timestamp = timestamp

        new_id = ((timestamp - TWEPOCH) << TIMESTAMP_LEFT_SHIFT) | (self.datacenter_id << DATACENTER_ID_SHIFT) | \
                 (self.worker_id << WOKER_ID_SHIFT) | self.sequence
        return new_id

    def _til_next_millis(self, last_timestamp):
        """
        等到下一毫秒
        """
        timestamp = self._gen_timestamp()
        while timestamp <= last_timestamp:
            timestamp = self._gen_timestamp()
        return timestamp


'''
===============所有方法===================
    element是查找一个标签
    elements是查找所有标签

    1、find_element_by_link_text  通过链接文本去找
    2、find_element_by_id 通过id去找
    3、find_element_by_class_name
    4、find_element_by_partial_link_text
    5、find_element_by_name
    6、find_element_by_css_selector
    7、find_element_by_tag_name
'''
driver = webdriver.Chrome()

try:
    # 发送请求
    url = 'http://www.minigps.net/fc.html'
    driver.get(url)
    driver.implicitly_wait(10)
    driver.maximize_window()  # 最大化

    readbook = openpyxl.load_workbook('C:\\Users\\administor\\Documents\\zb.xlsx')
    # 名字的方式
    sheetnames = readbook.get_sheet_names()  # 获取读文件中所有的sheet，通过名字的方式
    ws = readbook.get_sheet_by_name(sheetnames[0])  # 获取第一个sheet内容
    # 获取sheet的最大行数和列数
    rows = ws.max_row
    cols = ws.max_column + 1
    name = 'None'
    local = '['
    for r in range(1, rows):
        if (r > 1):
            local = local + ','
        local = local + '['
        for c in range(1, cols):
            if ('None' != str(ws.cell(r, 1).value) and 1 == c):
                local = local[0:-2]
                if (local != ""):
                    local = local + ']'
                    name = seaName
                    seaNumber = name
                    save(seaNumber, 2, 0, 0, 1, 2, local,
                         "", 1, '#00C395', 1, "", "", "")
                    print(name + "      " + local)
                local = '[['
                seaName = ws.cell(r, 1).value
                seaNumber = seaName

            elif (1 != c):
                value = ws.cell(r, c).value
                sb = str(value).split('°', 1)
                sb1 = sb[1].split('′', 1)
                value1 = sb[0]
                value2 = sb1[0]
                value3 = sb1[1]
                value3 = value3[0:-1]

                dfm1 = driver.find_element_by_id('input_dfm1')
                dfm1.clear()
                time.sleep(1)

                dfm1.send_keys(value1)
                time.sleep(1)

                # 5、find_element_by_id 通过id属性名查找
                dfm2 = driver.find_element_by_id('input_dfm2')
                dfm2.clear()
                time.sleep(1)

                dfm2.send_keys(value2)
                time.sleep(1)

                # 5、find_element_by_id 通过id属性名查找
                dfm3 = driver.find_element_by_id('input_dfm3')
                dfm3.clear()
                time.sleep(1)

                dfm3.send_keys(value3)
                time.sleep(1)
                driver.find_element_by_xpath('//*[@class="info3"]/input[4]').click()

                result = driver.find_element_by_id('calculated_du').text

                local = local + result
                if (2 == c):
                    local = local + ','

        local = local + ']'

    print("结束")
    while (1):
        time.sleep(100)
        print("清退出淘宝账号")



finally:
    driver.close()
